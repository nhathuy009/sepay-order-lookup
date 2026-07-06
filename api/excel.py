"""Serverless function: xử lý Excel hàng loạt.

Nhận POST JSON {file_base64, access_token}, đọc cột A (mã đơn), tra cứu từng
dòng và trả về file Excel (base64) đã điền kết quả cùng thống kê.
"""
import base64
import io
import json
import os
import sys
from http.server import BaseHTTPRequestHandler

import openpyxl

sys.path.append(os.path.dirname(__file__))
from _core import (  # noqa: E402
    lookup_order,
    check_access,
    detect_system,
    EXCEL_HEADERS,
    EXCEL_FIELDS,
)


class handler(BaseHTTPRequestHandler):
    def _send(self, status, payload):
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def do_POST(self):
        try:
            length = int(self.headers.get("Content-Length", 0))
            raw = self.rfile.read(length) if length else b"{}"
            body = json.loads(raw.decode("utf-8") or "{}")
        except (ValueError, json.JSONDecodeError):
            self._send(400, {"error": "JSON không hợp lệ"})
            return

        token = body.get("access_token") or self.headers.get("X-Access-Token", "")
        if not check_access(token):
            self._send(401, {"error": "Sai mật khẩu truy cập"})
            return

        file_b64 = body.get("file_base64", "")
        if not file_b64:
            self._send(400, {"error": "Thiếu file Excel"})
            return

        try:
            if "," in file_b64:  # bỏ tiền tố data:...;base64,
                file_b64 = file_b64.split(",", 1)[1]
            file_bytes = base64.b64decode(file_b64)
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
            sheet = wb.active
        except Exception as e:
            self._send(400, {"error": f"Không đọc được file Excel: {e}"})
            return

        for col_idx, header in enumerate(EXCEL_HEADERS, start=1):
            sheet.cell(row=1, column=col_idx, value=header)

        status_col = len(EXCEL_HEADERS)
        max_row = sheet.max_row
        total = 0
        success = 0

        for row in range(2, max_row + 1):
            cell_val = sheet.cell(row=row, column=1).value
            if not cell_val:
                continue
            total += 1
            order_code = str(cell_val).strip()

            if detect_system(order_code) is None:
                sheet.cell(row=row, column=status_col, value="Bỏ qua (Mã không hợp lệ)")
                continue

            result = lookup_order(order_code)
            for col_idx, field_name in enumerate(EXCEL_FIELDS, start=1):
                if field_name is not None:
                    sheet.cell(row=row, column=col_idx, value=result.get(field_name, ""))
            if result.get("status_msg") == "Thành công":
                success += 1

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        result_b64 = base64.b64encode(out.read()).decode("ascii")

        self._send(200, {
            "file_base64": result_b64,
            "total": total,
            "success": success,
        })
