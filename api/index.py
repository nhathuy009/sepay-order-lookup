"""Serverless function duy nhất cho toàn bộ API."""
import base64
import io
import json
import os
import sys
import urllib.parse
import requests
import re
from http.server import BaseHTTPRequestHandler

import openpyxl
from calendar import monthrange
from datetime import date, datetime, timedelta

sys.path.append(os.path.dirname(__file__))
from _core import (  # noqa: E402
    lookup_order,
    lookup_customer,
    check_access,
    detect_system,
    reverse_lookup_orders_by_invoices,
    APP_ACCESS_TOKEN,
    EXCEL_HEADERS,
    EXCEL_FIELDS,
)

from collections import defaultdict
# Thêm dòng này vào cụm import từ file nội bộ
from _payment import search_sepay_transaction, list_sepay_transactions, get_sepay_bank_accounts
from _invoice import lookup_invoice, fetch_invoices_by_date
from _gdt_invoice import lookup_gdt_invoices, lookup_gdt_invoices_by_type, gdt_fetch_invoice_detail
from _invoiceBKAV import ehoadon_login, ehoadon_buyer_search, ehoadon_invoice_create, ehoadon_invoice_list

def handle_fetch_employees_excel(body):
    file_b64 = body.get("file_base64", "")
    if not file_b64:
        return 400, {"error": "Thiếu file Excel"}
    try:
        # Tách phần header của base64 nếu có
        if "," in file_b64:
            file_b64 = file_b64.split(",", 1)[1]
        
        # Đọc dữ liệu excel trong RAM bằng openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(base64.b64decode(file_b64)), data_only=True)
        
        # Biểu thức chính quy (Regex) bắt các sheet có tên như T012026, T122026...
        pattern = re.compile(r"^T\d{2}20\d{2}$")
        sheets_data = {}
        
        for sheet_name in wb.sheetnames:
            clean_name = sheet_name.strip()
            if pattern.match(clean_name):
                ws = wb[sheet_name]
                data = []
                # Lấy dữ liệu từ dòng 7 trở đi
                for row in range(7, ws.max_row + 1):
                    ma_nv = ws.cell(row=row, column=2).value  # Cột B
                    ten_nv = ws.cell(row=row, column=3).value # Cột C
                    
                    # Nếu dòng có chứa mã NV hoặc Tên NV thì mới đưa vào danh sách
                    if ma_nv is not None or ten_nv is not None:
                        data.append({
                            "ma_nv": str(ma_nv).strip() if ma_nv is not None else "",
                            "ten_nv": str(ten_nv).strip() if ten_nv is not None else ""
                        })
                
                # Lưu mảng data theo key là tên sheet
                sheets_data[clean_name] = data
                
        return 200, {"sheets": sheets_data}
        
    except Exception as e:
        return 400, {"error": f"Lỗi đọc file Excel: {str(e)}"}
        
def handle_invoice(body):
    code = (body.get("code") or "").strip()
    if not code:
        return 400, {"error": "Thiếu số hóa đơn"}
    return 200, lookup_invoice(code)

def _dmy_to_iso_date(d):
    """Đổi 'DD/MM/YYYY' (định dạng flatpickr đang dùng ở frontend) sang 'YYYY-MM-DD'
    (định dạng dateFrom/dateTo mà API admin/orders của 10X & Solobiz yêu cầu)."""
    d = (d or "").strip()
    parts = d.split("/")
    if len(parts) == 3:
        dd, mm, yyyy = parts
        return f"{yyyy}-{mm.zfill(2)}-{dd.zfill(2)}"
    return d

def _month_range_iso(dmy_date):
    """Từ 1 ngày "dd/mm/yyyy" -> (date_from, date_to) bao trọn THÁNG chứa ngày đó,
    định dạng "YYYY-MM-DD". Nới thêm 1 ngày đầu/cuối tháng để tránh lệch múi
    giờ/ngày ghi nhận giữa hệ thống hóa đơn và hệ thống đơn hàng (giống khuyến
    nghị trong docstring của reverse_lookup_orders_by_invoices). Trả về
    (None, None) nếu không parse được (ProcessInvNote không có phần ngày).
    """
    dmy_date = (dmy_date or "").strip()
    parts = dmy_date.split("/")
    if len(parts) != 3:
        return None, None
    try:
        dd, mm, yyyy = int(parts[0]), int(parts[1]), int(parts[2])
        first_day = date(yyyy, mm, 1)
        last_day = date(yyyy, mm, monthrange(yyyy, mm)[1])
    except (ValueError, TypeError):
        return None, None
    date_from = first_day - timedelta(days=1)
    date_to = last_day + timedelta(days=1)
    return date_from.isoformat(), date_to.isoformat()


def _strip_course_duration_suffix(title):
    """Bỏ hậu tố kỳ hạn dạng '- N tháng' ở cuối tên khóa học, để gộp các kỳ hạn
    (3/6/12 tháng...) của cùng 1 khóa học vào chung 1 nhóm/1 bảng.

    VD: "MEMBERSHIP-10xVIP (Thầy Hùng Bigman) - 3 tháng"
        -> "MEMBERSHIP-10xVIP (Thầy Hùng Bigman)"
    Nếu tên không khớp mẫu (không có hậu tố kỳ hạn), trả về nguyên văn.
    """
    title = (title or "").strip()
    return re.sub(r"\s*-\s*\d+\s*th[áa]ng\s*$", "", title, flags=re.IGNORECASE).strip() or title

def handle_invoice_by_date(body):
    start_date = (body.get("start_date") or "").strip()
    end_date = (body.get("end_date") or "").strip()
    invoice_kind = (body.get("invoice_kind") or "2").strip()

    if not start_date or not end_date:
        return 400, {"error": "Vui lòng chọn Từ ngày và Đến ngày."}

    result = fetch_invoices_by_date(start_date, end_date, invoice_kind)
    if result is None:
        return 400, {"error": "Không lấy được dữ liệu hóa đơn (lỗi đăng nhập hoặc kết nối tới hệ thống hóa đơn)."}

    # Tra ngược mã đơn hàng (10X/Solobiz) cho từng số hóa đơn vừa lấy được từ SePay eInvoice.
    invoice_numbers = [inv.get("invoice_no", "") for inv in result if inv.get("invoice_no")]
    if invoice_numbers:
        order_map = reverse_lookup_orders_by_invoices(
            invoice_numbers,
            date_from=_dmy_to_iso_date(start_date),
            date_to=_dmy_to_iso_date(end_date),
        )
        for inv in result:
            match = order_map.get(inv.get("invoice_no", ""), {})
            inv["order_code"] = match.get("order_code", "")
            inv["order_system"] = match.get("system", "")
            inv["lead_name"] = match.get("lead_name", "")
            inv["username"] = match.get("username", "")
            inv["ref_username"] = match.get("ref_username", "")
            inv["commission_rate"] = match.get("commission_rate", "")
            inv["hoahong"] = match.get("hoahong", "")
            inv["item_id"] = match.get("item_id", "")
            inv["item_title"] = match.get("item_title", "")

    # Hóa đơn số tiền âm có "adjusts_invoice_no" (trích từ ProcessInvNote ở
    # _invoice.py) là số hóa đơn GỐC bị điều chỉnh/thay thế. Tra trong CHÍNH danh
    # sách hóa đơn vừa lấy được (cùng đợt/khoảng ngày) - nếu thấy, ghi chú vào cột
    # Note của DÒNG HÓA ĐƠN GỐC (không phải dòng điều chỉnh/thay thế), phân biệt 2 loại:
    #   - "dieu_chinh" (điều chỉnh giảm, thường là hoàn tiền): tiền hóa đơn gốc
    #     không còn giá trị thật -> frontend sẽ xóa trắng 7 cột tiền/ref của dòng gốc.
    #   - "thay_the" (thay thế, VD đổi cá nhân sang công ty): tiền hóa đơn gốc vẫn
    #     là tiền thật -> giữ nguyên 7 cột, chỉ thêm ghi chú.
    invoice_by_no = {inv.get("invoice_no", ""): inv for inv in result if inv.get("invoice_no")}
    for inv in result:
        adjusts_no = inv.get("adjusts_invoice_no", "")
        if not adjusts_no:
            continue
        original = invoice_by_no.get(adjusts_no)
        if original is None:
            continue
        adj_type = inv.get("adjustment_type", "")
        inv_no = inv.get("invoice_no", "")
        if adj_type == "thay_the":
            original["note"] = f"KH đổi sang cty, HĐ thay thế {inv_no}"
            original["note_type"] = "thay_the"
        else:
            # Mặc định coi là điều chỉnh giảm (giữ hành vi cũ) nếu ProcessInvNote
            # không khớp mẫu "Thay thế..." đã biết.
            original["note"] = f"KH hoàn tiền, HĐ điều chỉnh {inv_no}"
            original["note_type"] = "dieu_chinh"

    # Hóa đơn điều chỉnh/thay thế (có adjusts_invoice_no) mà bản thân KHÔNG tra
    # ngược được đơn hàng (order_code rỗng) -> lấy thông tin đơn hàng từ HÓA ĐƠN
    # GỐC bị điều chỉnh, chia làm 2 bước:
    #   B1 (trong batch): hóa đơn gốc nằm CÙNG đợt/khoảng ngày đang tra -> đã có
    #      sẵn trong invoice_by_no, dùng lại luôn, không tốn thêm lượt gọi API.
    #   B2 (ngoài batch): hóa đơn gốc KHÔNG nằm trong khoảng ngày đang tra (VD tra
    #      tháng 8 nhưng hóa đơn gốc phát hành tháng 6) -> gom các "adjusts_no"
    #      còn thiếu lại, gọi 1 LẦN reverse_lookup_orders_by_invoices KHÔNG giới
    #      hạn date_from/date_to (tìm trong toàn bộ lịch sử đơn hàng) để tra ngược
    #      trực tiếp theo đúng số hóa đơn gốc đó.
    # Nhờ vậy hóa đơn điều chỉnh/thay thế thoát khỏi nhóm "Chưa xác định" và được
    # xếp đúng vào bảng khóa học của hóa đơn gốc.
    def _to_negative(value):
        """Ép 1 giá trị số về ÂM (giữ nguyên nếu không parse được thành số).
        Dùng cho các cột tiền/% Ref copy từ hóa đơn/đơn hàng GỐC (vốn dương) sang
        dòng hóa đơn điều chỉnh/thay thế (vốn mang ý nghĩa hoàn/trừ tiền)."""
        if value in ("", None):
            return value
        try:
            num = float(str(value).replace(",", "").strip())
        except (TypeError, ValueError):
            return value
        return -abs(num)

    REVERSE_MATCHED_COMMISSION_RATE = 25  # % REF cố định cho hóa đơn reverse_matched

    def _copy_order_info(dest, src, adjusts_no):
        dest["order_code"] = src.get("order_code", "")
        dest["order_system"] = src.get("order_system") or src.get("system", "")
        dest["lead_name"] = src.get("lead_name", "")
        dest["username"] = src.get("username", "")
        dest["ref_username"] = src.get("ref_username", "")
        dest["item_id"] = src.get("item_id", "")
        dest["item_title"] = src.get("item_title", "")
        existing_note = dest.get("note") or ""
        dest["note"] = (existing_note + " " if existing_note else "") + f"(Đơn hàng lấy từ HĐ gốc {adjusts_no})"
        # note_type riêng ("reverse_matched") để frontend phân biệt với "dieu_chinh"
        # (hóa đơn GỐC bị điều chỉnh giảm - cần xóa trắng cột) và "thay_the" - dòng
        # này KHÔNG được xóa trắng cột, KHÔNG bị loại khỏi tổng, vì tiền của nó
        # (đã ép âm ở dưới) là số thật cần cộng vào tổng để trừ đi phần đã hoàn.
        dest["note_type"] = "reverse_matched"
        # Đánh dấu để: (1) sắp dòng này xuống CUỐI bảng khi sort ở bước gộp
        # nhóm khóa học bên dưới, (2) ép cột tiền của CHÍNH hóa đơn về ÂM, vì đây
        # là hóa đơn điều chỉnh/thay thế (bản chất là hoàn/trừ tiền so với hóa đơn
        # gốc), không nên hiển thị dương như đơn hàng gốc.
        dest["is_reverse_matched"] = True
        for money_field in ("amount_before_tax", "vat_amount", "total_amount"):
            dest[money_field] = _to_negative(dest.get(money_field))
        # % REF cố định 25% cho MỌI hóa đơn reverse_matched (không lấy commission_rate
        # copy từ đơn hàng gốc nữa - vì đơn gốc có thể mang % hoa hồng khác, nhưng
        # quy ước hoàn/trừ tiền REF luôn tính theo mức cố định này).
        dest["commission_rate"] = REVERSE_MATCHED_COMMISSION_RATE
        # Số tiền REF = Số tiền trước thuế (đã âm ở trên) x % REF -> tự ra số âm.
        # Làm tròn về SỐ NGUYÊN (không thập phân) giống mọi giá trị "hoahong" khác
        # trong hệ thống (vốn luôn là số nguyên VNĐ lấy thẳng từ đơn hàng) - tránh
        # hiển thị lẻ kiểu "-1.041.666,75" do phép nhân phần trăm sinh ra.
        try:
            amount_num = float(dest.get("amount_before_tax") or 0)
        except (TypeError, ValueError):
            amount_num = 0.0
        dest["hoahong"] = round(amount_num * (REVERSE_MATCHED_COMMISSION_RATE / 100))

    still_missing = []  # list[(inv, adjusts_no)] cần tra ngược ngoài batch
    for inv in result:
        if inv.get("order_code"):
            continue
        adjusts_no = inv.get("adjusts_invoice_no", "")
        if not adjusts_no:
            continue
        original = invoice_by_no.get(adjusts_no)
        if original is not None and original.get("order_code"):
            _copy_order_info(inv, original, adjusts_no)
        else:
            still_missing.append((inv, adjusts_no))

    if still_missing:
        # Gom theo THÁNG của hóa đơn gốc (lấy từ "adjusts_invoice_date" - do
        # _invoice.py trích ra cùng với adjusts_invoice_no, dạng "dd/mm/yyyy")
        # để mỗi lượt gọi reverse_lookup_orders_by_invoices chỉ quét đúng
        # khoảng THÁNG đó thay vì toàn bộ lịch sử đơn hàng - nhẹ hơn hẳn.
        # Hóa đơn nào không trích được ngày (ProcessInvNote không có "ngày...")
        # đành gộp vào nhóm tra không giới hạn ngày (hành vi dự phòng cũ).
        month_groups = {}  # "mm/yyyy" -> {"date_from", "date_to", "items": [(inv, adjusts_no)]}
        no_date_items = []
        for inv, adjusts_no in still_missing:
            adjusts_date = inv.get("adjusts_invoice_date", "")
            date_from, date_to = _month_range_iso(adjusts_date)
            if date_from is None:
                no_date_items.append((inv, adjusts_no))
                continue
            month_key = adjusts_date[3:]  # "mm/yyyy"
            group = month_groups.setdefault(
                month_key, {"date_from": date_from, "date_to": date_to, "items": []}
            )
            group["items"].append((inv, adjusts_no))

        for group in month_groups.values():
            extra_numbers = sorted({adjusts_no for _, adjusts_no in group["items"]})
            extra_order_map = reverse_lookup_orders_by_invoices(
                extra_numbers, date_from=group["date_from"], date_to=group["date_to"]
            )
            for inv, adjusts_no in group["items"]:
                match = extra_order_map.get(adjusts_no)
                if match and match.get("order_code"):
                    _copy_order_info(inv, match, adjusts_no)

        if no_date_items:
            extra_numbers = sorted({adjusts_no for _, adjusts_no in no_date_items})
            extra_order_map = reverse_lookup_orders_by_invoices(extra_numbers)  # không giới hạn ngày
            for inv, adjusts_no in no_date_items:
                match = extra_order_map.get(adjusts_no)
                if match and match.get("order_code"):
                    _copy_order_info(inv, match, adjusts_no)

    # Mã ĐH (order_code) phải là DUY NHẤT trong bảng kết quả. LUÔN ưu tiên giữ
    # hóa đơn KHÔNG PHẢI reverse_matched (hóa đơn gốc/khớp trực tiếp) khi trùng
    # Mã ĐH với 1 hóa đơn reverse_matched - loại bỏ HOÀN TOÀN hóa đơn reverse_matched
    # đó khỏi kết quả. LƯU Ý: không thể dựa vào thứ tự "gặp trước - gặp sau" khi
    # duyệt result, vì SePay eInvoice trả về MỚI NHẤT TRƯỚC - mà hóa đơn
    # reverse_matched (điều chỉnh/thay thế) luôn phát sinh SAU hóa đơn gốc theo
    # thời gian, nên trong danh sách thô nó thường đứng TRƯỚC hóa đơn gốc. Vì vậy
    # phải quét 2 lượt: lượt 1 xác định trước các Mã ĐH đã có ở hóa đơn KHÔNG phải
    # reverse_matched, lượt 2 mới lọc bỏ.
    order_codes_on_non_reverse = {
        inv.get("order_code", "")
        for inv in result
        if inv.get("order_code") and inv.get("note_type") != "reverse_matched"
    }
    seen_reverse_matched_codes = set()
    deduped_result = []
    for inv in result:
        order_code = inv.get("order_code", "")
        if order_code and inv.get("note_type") == "reverse_matched":
            if order_code in order_codes_on_non_reverse:
                continue  # đã có hóa đơn gốc/khớp trực tiếp cùng Mã ĐH -> bỏ dòng này
            if order_code in seen_reverse_matched_codes:
                continue  # 2 hóa đơn reverse_matched cùng Mã ĐH -> chỉ giữ dòng đầu tiên
            seen_reverse_matched_codes.add(order_code)
        deduped_result.append(inv)
    result = deduped_result

    # Phân loại hàng hóa: mỗi item.id là 1 khóa học/kỳ hạn khác nhau. Một số khóa
    # học có nhiều kỳ hạn con (VD "... - 3 tháng", "... - 6 tháng", "... - 12 tháng")
    # nhưng vẫn nên gộp chung 1 bảng/1 tab - nên gộp nhóm theo TÊN ĐÃ CHUẨN HÓA
    # (bỏ hậu tố kỳ hạn) thay vì theo item_id thô. Đơn không tra ngược được
    # (chưa rõ khóa học nào) gộp vào nhóm "Chưa xác định".
    courses = {}
    for inv in result:
        item_title = inv.get("item_title") or "Chưa xác định (không tra ngược được đơn hàng)"
        group_key = _strip_course_duration_suffix(item_title)
        if group_key not in courses:
            courses[group_key] = {"title": group_key, "invoices": []}
        courses[group_key]["invoices"].append(inv)

    # SePay eInvoice trả về mặc định mới nhất trước - đổi lại thành cũ -> mới
    # theo "arising_date" (định dạng "YYYY-MM-DD" nên so sánh chuỗi trực tiếp
    # là đủ, không cần parse ngày). Hóa đơn thiếu ngày (hiếm khi xảy ra) đẩy
    # xuống cuối bảng thay vì lên đầu.
    # Riêng các hóa đơn điều chỉnh/thay thế được tra ngược thông tin đơn hàng từ
    # hóa đơn GỐC (is_reverse_matched=True, xem _copy_order_info ở trên) luôn bị
    # đẩy xuống DƯỚI CÙNG bảng, sau tất cả hóa đơn khớp trực tiếp - dù ngày phát
    # hành của chúng là gì.
    for group in courses.values():
        group["invoices"].sort(
            key=lambda inv: (
                1 if inv.get("is_reverse_matched") else 0,
                inv.get("arising_date") or "9999-99-99",
            )
        )

    return 200, {"courses": list(courses.values()), "total": len(result)}

def handle_gdt_invoice(body):
    username = (body.get("username") or "").strip()
    password = body.get("password") or ""
    start_date = (body.get("start_date") or "").strip()
    end_date = (body.get("end_date") or "").strip()
    is_purchase = bool(body.get("is_purchase", True))

    if not username or not password:
        return 400, {"error": "Vui lòng nhập Mã số thuế và Mật khẩu."}
    if not start_date or not end_date:
        return 400, {"error": "Vui lòng chọn Từ ngày và Đến ngày."}

    res = lookup_gdt_invoices(username, password, start_date, end_date, is_purchase)
    status = 400 if "error" in res else 200
    return status, res

def handle_gdt_invoice_by_type(body):
    username = (body.get("username") or "").strip()
    password = body.get("password") or ""
    start_date = (body.get("start_date") or "").strip()
    end_date = (body.get("end_date") or "").strip()
    is_purchase = bool(body.get("is_purchase", True))
    token = (body.get("token") or "").strip() or None

    if not username or not password:
        return 400, {"error": "Vui lòng nhập Mã số thuế và Mật khẩu."}
    if not start_date or not end_date:
        return 400, {"error": "Vui lòng chọn Từ ngày và Đến ngày."}

    try:
        ttxly = int(body.get("ttxly"))
    except (TypeError, ValueError):
        return 400, {"error": "Thiếu hoặc sai tham số ttxly (chỉ chấp nhận 5, 6 hoặc 8)."}

    res = lookup_gdt_invoices_by_type(username, password, start_date, end_date, is_purchase, ttxly, token=token)
    status = 400 if "error" in res else 200
    return status, res

def handle_gdt_invoice_detail(body):
    username = (body.get("username") or "").strip()
    password = body.get("password") or ""
    invoice = body.get("invoice") or {}

    if not username or not password:
        return 400, {"error": "Vui lòng nhập Mã số thuế và Mật khẩu."}
    if not isinstance(invoice, dict) or not invoice:
        return 400, {"error": "Thiếu thông tin hóa đơn cần tra cứu chi tiết."}

    res = gdt_fetch_invoice_detail(username, password, invoice)
    status = 400 if "error" in res else 200
    return status, res

def handle_lookup(body):
    code = (body.get("code") or "").strip()
    if not code:
        return 400, {"error": "Thiếu mã đơn hàng / mã KH"}
    sa_system = (body.get("sa_system") or "").strip() or None
    return 200, lookup_customer(code, sa_system=sa_system)
    
def handle_excel(body):
    file_b64 = body.get("file_base64", "")
    if not file_b64:
        return 400, {"error": "Thiếu file Excel"}
    try:
        if "," in file_b64:
            file_b64 = file_b64.split(",", 1)[1]
        wb = openpyxl.load_workbook(io.BytesIO(base64.b64decode(file_b64)))
        sheet = wb.active
    except Exception as e:
        return 400, {"error": f"Không đọc được file Excel: {e}"}

    for col_idx, header in enumerate(EXCEL_HEADERS, start=1):
        sheet.cell(row=1, column=col_idx, value=header)

    status_col = len(EXCEL_HEADERS)
    total = 0
    success = 0
    for row in range(2, sheet.max_row + 1):
        cell_val = sheet.cell(row=row, column=1).value
        if not cell_val:
            continue
        total += 1
        order_code = str(cell_val).strip()
        if detect_system(order_code) is None:
            sheet.cell(row=row, column=status_col, value="Bỏ qua (Mã không hợp lệ)")
            continue
        result = lookup_order(order_code)
        for col_idx, field_name in enumerate(EXCEL_FIELDS, start=1):
            if field_name is not None:
                sheet.cell(row=row, column=col_idx, value=result.get(field_name, ""))
        if result.get("status_msg") == "Thành công":
            success += 1

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return 200, {
        "file_base64": base64.b64encode(out.read()).decode("ascii"),
        "total": total,
        "success": success,
    }
    
def handle_bank_statement(body):
    file_b64 = body.get("file_base64", "")
    if not file_b64:
        return 400, {"error": "Thiếu file Excel"}
    try:
        if "," in file_b64:
            file_b64 = file_b64.split(",", 1)[1]
        wb = openpyxl.load_workbook(io.BytesIO(base64.b64decode(file_b64)), data_only=True)
        sheet = wb.active
    except Exception as e:
        return 400, {"error": f"Không đọc được file Excel: {e}"}

    so_du_dau_ky_raw = sheet.cell(row=7, column=4).value
    so_du_dau_ky = 0
    if so_du_dau_ky_raw is not None:
        try:
            chuoi_so = str(so_du_dau_ky_raw).replace(",", "").replace(" ", "").strip()
            so_du_dau_ky = float(chuoi_so)
        except ValueError:
            so_du_dau_ky = 0

    # Lấy số tài khoản từ ô A4, tương tự hàm RIGHT(A4, 8) trong Excel
    so_tk_raw = sheet.cell(row=4, column=1).value
    so_tk_str = str(so_tk_raw).strip() if so_tk_raw is not None else ""
    so_tk_8_ky_tu = so_tk_str[-8:] if so_tk_str else ""
    nhan_dau_ky = f"Số dư đầu kỳ STK {so_tk_8_ky_tu}" if so_tk_8_ky_tu else "Số dư đầu kỳ STK ..."

    dem_gui_vao = defaultdict(int)
    dem_rut_ra = defaultdict(int)

    current_row = 9
    last_valid_row = 8 

    while True:
        rut_ra = sheet.cell(row=current_row, column=5).value
        gui_vao = sheet.cell(row=current_row, column=6).value
        ngay_gd = sheet.cell(row=current_row, column=1).value 

        if ngay_gd is None and rut_ra is None and gui_vao is None:
            break

        last_valid_row = current_row 

        if rut_ra is not None:
            try:
                val = float(rut_ra)
                if val > 0: dem_rut_ra[val] += 1
            except ValueError: pass 

        if gui_vao is not None:
            try:
                val = float(gui_vao)
                if val > 0: dem_gui_vao[val] += 1
            except ValueError: pass

        current_row += 1

    so_du_cuoi_ky_raw = sheet.cell(row=last_valid_row, column=7).value
    so_du_cuoi_ky = 0
    if so_du_cuoi_ky_raw is not None:
        try:
            chuoi_so_cuoi = str(so_du_cuoi_ky_raw).replace(",", "").replace(" ", "").strip()
            so_du_cuoi_ky = float(chuoi_so_cuoi)
        except ValueError:
            so_du_cuoi_ky = 0

    tat_ca_gia_tri = set(dem_gui_vao.keys()).union(set(dem_rut_ra.keys()))

    def sort_key_ghi_chu(gia_tri):
        if gia_tri < 50000:
            return (1, gia_tri)      # Nhóm số nhỏ (< 50.000) -> xếp dưới
        elif gia_tri > 11000000:
            return (2, gia_tri)      # Nhóm số lớn (> 11.000.000) -> xếp dưới cùng
        else:
            return (0, gia_tri)      # Nhóm giữa (50.000 - 11.000.000) -> xếp trên, tăng dần

    tat_ca_gia_tri = sorted(tat_ca_gia_tri, key=sort_key_ghi_chu)

    wb_new = openpyxl.Workbook()
    ws = wb_new.active
    ws.title = "Tong_Hop_Sao_Ke"

    ws.append(["Nội dung diễn giải", "Gửi vào", "Rút ra", "Số dư lũy kế", "Ghi chú (Để bạn dò số)"])
    ws.append([nhan_dau_ky, "", "", so_du_dau_ky, ""])

    current_excel_row = 3
    for gia_tri in tat_ca_gia_tri:
        sl_vao = dem_gui_vao[gia_tri]
        sl_ra = dem_rut_ra[gia_tri]
        
        tong_vao = (gia_tri * sl_vao) if sl_vao > 0 else ""
        tong_ra = -(gia_tri * sl_ra) if sl_ra > 0 else "" 
        
        ghi_chu = []
        if sl_vao > 0: ghi_chu.append(f"Vào: {gia_tri:,.0f} x {sl_vao}")
        if sl_ra > 0: ghi_chu.append(f"Ra: {gia_tri:,.0f} x {sl_ra}")
        
        ws.append([
            "", 
            tong_vao, 
            tong_ra, 
            f"=D{current_excel_row-1}+SUM(B{current_excel_row}:C{current_excel_row})", 
            " | ".join(ghi_chu)
        ])
        current_excel_row += 1

    ws.append([
        "Số dư cuối kỳ", 
        f"=SUM(B3:B{current_excel_row-1})", 
        f"=SUM(C3:C{current_excel_row-1})", 
        so_du_cuoi_ky, 
        ""
    ])

    # ==========================================
    # LÀM ĐẸP GIAO DIỆN EXCEL (CẬP NHẬT TẠI ĐÂY)
    # ==========================================
    # Ép format phân cách phần ngàn, không thập phân cho toàn bộ cột B, C, D (Kể cả ô chứa công thức)
    for row in ws.iter_rows(min_row=2, max_row=current_excel_row, min_col=2, max_col=4):
        for cell in row:
            # '#,##0' là format chuẩn: có dấu phẩy phần ngàn, không có số sau dấu chấm thập phân
            # ;(#,##0) đảm bảo nếu là số âm thì tự chui vào trong ngoặc đơn
            cell.number_format = '#,##0;(#,##0)'
    
    for col in range(1, 6):
        ws.cell(row=1, column=col).font = openpyxl.styles.Font(bold=True)
        ws.cell(row=current_excel_row, column=col).font = openpyxl.styles.Font(bold=True)
    ws.cell(row=2, column=1).font = openpyxl.styles.Font(bold=True) 

    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 45

    out = io.BytesIO()
    wb_new.save(out)
    out.seek(0)
    return 200, {
        "file_base64": base64.b64encode(out.read()).decode("ascii")
    } 
    
# ==========================================
# EHOADON (van.ehoadon.vn) - TẠO HÓA ĐƠN TỰ ĐỘNG
# ==========================================
# Toàn bộ logic nghiệp vụ (đăng nhập, tìm KH, tạo hóa đơn, lấy danh sách)
# nằm trong _invoiceBKAV.py. Ở đây chỉ validate request body + quyết định
# HTTP status code, giống hệt cách các action khác (vd. handle_gdt_invoice)
# đang điều hướng sang _gdt_invoice.py.
def handle_ehoadon_login(body):
    username = (body.get("username") or "").strip()
    password = body.get("password") or ""
    if not username or not password:
        return 400, {"error": "Thiếu username hoặc password eHoadon"}
    res = ehoadon_login(username, password)
    return (400 if "error" in res else 200), res


def handle_ehoadon_buyer_search(body):
    cookies = body.get("cookies") or {}
    keyword = (body.get("keyword") or "").strip()
    if not cookies:
        return 400, {"error": "Thiếu phiên đăng nhập eHoadon (cookies). Vui lòng đăng nhập lại."}
    res = ehoadon_buyer_search(cookies, keyword)
    return (400 if "error" in res else 200), res


def handle_ehoadon_invoice_create(body):
    cookies = body.get("cookies") or {}
    buyer_info = body.get("buyer_info") or {}
    note_input = body.get("note") or "Hóa đơn tự động"
    items = body.get("items") or []
    if not cookies:
        return 400, {"error": "Thiếu phiên đăng nhập eHoadon (cookies). Vui lòng đăng nhập lại."}
    if not items:
        return 400, {"error": "Phải có ít nhất 1 hàng hóa"}
    res = ehoadon_invoice_create(cookies, buyer_info, note_input, items)
    return (400 if "error" in res else 200), res


def handle_ehoadon_invoice_list(body):
    cookies = body.get("cookies") or {}
    if not cookies:
        return 400, {"error": "Thiếu phiên đăng nhập eHoadon (cookies). Vui lòng đăng nhập lại."}
    today_str = datetime.now().strftime("%Y-%m-%d")
    from_date = (body.get("from_date") or "").strip() or today_str
    to_date = (body.get("to_date") or "").strip() or today_str
    res = ehoadon_invoice_list(cookies, from_date, to_date)
    return (400 if "error" in res else 200), res


class handler(BaseHTTPRequestHandler):
    def _send(self, status, payload):
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def do_GET(self):
        self._send(200, {"auth_required": bool(APP_ACCESS_TOKEN)})

    def do_POST(self):
        try:
            length = int(self.headers.get("Content-Length", 0))
            raw = self.rfile.read(length) if length else b"{}"
            body = json.loads(raw.decode("utf-8") or "{}")
        except (ValueError, json.JSONDecodeError):
            self._send(400, {"error": "JSON không hợp lệ"})
            return

        token = body.get("access_token") or self.headers.get("X-Access-Token", "")
        if not check_access(token):
            self._send(401, {"error": "Sai mật khẩu truy cập"})
            return

        action = body.get("action", "lookup")
        
        if action == "excel":
            status, payload = handle_excel(body)
        elif action == "bank_statement":
            status, payload = handle_bank_statement(body)
        elif action == "lookup":
            status, payload = handle_lookup(body)
        elif action == "invoice":
            status, payload = handle_invoice(body)
        elif action == "invoice_by_date":
            status, payload = handle_invoice_by_date(body)
        elif action == "gdt_invoice":
            status, payload = handle_gdt_invoice(body)
        elif action == "gdt_invoice_by_type":
            status, payload = handle_gdt_invoice_by_type(body)
        elif action == "gdt_invoice_detail":
            status, payload = handle_gdt_invoice_detail(body)
        elif action == "fetch_employees_excel":
            status, payload = handle_fetch_employees_excel(body)
        elif action == "search_transaction":
            code = (body.get("code") or "").strip()
            if not code:
                status, payload = 400, {"error": "Thiếu mã đơn hàng"}
            else:
                res = search_sepay_transaction(code)
                status, payload = (400 if "error" in res else 200), res
        elif action == "list_transactions":
            date_from = body.get("date_from", "").strip()
            date_to = body.get("date_to", "").strip()
            bank_brand = body.get("bank_brand", "").strip()
            bank_account = body.get("bank_account", "").strip()
            
            if not date_from or not date_to:
                status, payload = 400, {"error": "Thiếu thông tin ngày bắt đầu/kết thúc."}
            else:
                res = list_sepay_transactions(date_from, date_to, bank_brand, bank_account)
                status, payload = (400 if "error" in res else 200), res
        elif action == "get_bank_accounts":
            res = get_sepay_bank_accounts()
            status, payload = (400 if "error" in res else 200), res
        elif action == "ehoadon_login":
            status, payload = handle_ehoadon_login(body)
        elif action == "ehoadon_buyer_search":
            status, payload = handle_ehoadon_buyer_search(body)
        elif action == "ehoadon_invoice_create":
            status, payload = handle_ehoadon_invoice_create(body)
        elif action == "ehoadon_invoice_list":
            status, payload = handle_ehoadon_invoice_list(body)
        else:
            status, payload = 400, {"error": f"action không hợp lệ: {action}"}       
        self._send(status, payload)
